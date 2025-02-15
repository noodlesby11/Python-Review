#openyxl是用于处理excel文件的第三方库
import openpyxl
#创建一个excel工作簿
workbook=openpyxl.Workbook()
#在excel文件中创建工作表
sheet=workbook.create_sheet('示例文件')
#在工作表中添加数据
lis = ['A','B','C']
sheet.append(lis)
#保存文件
workbook.save('示例文件.xlsx')

#打开文件
wk=openpyxl.load_workbook('示例文件.xlsx')
#指定工作表
st=workbook['示例文件']
#表格是二维列表
lst=[]#存储行数据
for row in sheet.rows:
    sublst=[]#存储单元格数据
    for cell in row:#cell指单元格
        sublst.append(cell.value)
    lst.append(sublst)
for item in lst:
    print(item)